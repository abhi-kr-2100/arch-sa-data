from time import time
from tempfile import NamedTemporaryFile
from pickle import dumps, loads

from openpyxl import load_workbook

from django.views.generic.base import TemplateView
from django.views.generic.edit import FormView
from django.views.generic.list import ListView

from django.http import HttpResponse
from django.shortcuts import get_object_or_404

from .forms import TiliaTemplateUploadForm
from .models import LocationInformation, ValidationInformation, TiliaExcelFile


class HomePageView(TemplateView):
    template_name = "isodatabank/home.html"


class MoreAboutIsotopes(TemplateView):
    template_name = "isodatabank/know_more.html"


class RelatedWebsites(TemplateView):
    template_name = "isodatabank/related.html"

    
class UploadDataView(FormView):
    template_name = "isodatabank/upload_data.html"
    form_class = TiliaTemplateUploadForm
    success_url = '/upload/success/'

    def add_to_db(self, name, email, file):
        """Add uploaded data to isodatabank's database."""

        dataset_id = str(time())
        tilia_workbook = load_workbook(file)
        bg_info = tilia_workbook.worksheets[0]

        validation_info = ValidationInformation(
            submission_by=name, email=email,
            dataset_id=dataset_id
        )

        location_name = str(bg_info.cell(1, 2).value).strip()
        description = str(bg_info.cell(2, 2).value).strip()
        latitude = float(bg_info.cell(3, 2).value)
        longitude = float(bg_info.cell(4, 2).value)
        original_paper_url = str(bg_info.cell(5, 2).value).strip()
        archaeological_material_type = str(bg_info.cell(6, 2).value).strip()
        isotopes = str(bg_info.cell(7, 2).value).strip()

        location_info = LocationInformation(
            location_name=location_name, description=description,
            latitude=latitude, longitude=longitude,
            original_paper_url=original_paper_url,
            archaeological_material_type=archaeological_material_type,
            isotopes=isotopes,
            dataset_id=dataset_id
        )

        excel_data = TiliaExcelFile(data=dumps(tilia_workbook), dataset_id=dataset_id)

        excel_data.save()
        validation_info.save()
        location_info.save()

    def post(self, request, *args, **kwargs):
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        
        name = request.POST['name']
        email = request.POST['email']
        tilia_file = request.FILES['tilia_template_file']

        if form.is_valid():
            try:
                self.add_to_db(name, email, tilia_file)
            except Exception:
                return self.form_invalid(form)
                
            return self.form_valid(form)
        else:
            return self.form_invalid(form)


class UploadSuccessful(TemplateView):
    template_name = "isodatabank/upload_success.html"


class LocationListView(ListView):
    """Show a list of all available and validated locations.
    
    Locations are denoted by coordinates."""

    model = LocationInformation

    def get_template_names(self):
        return 'isodatabank/location_list.html'

    def get_context_data(self, **kwargs):
        validated = ValidationInformation.objects.filter(validated=True)
        locs = [
            l for v in validated \
                for l in LocationInformation.objects.filter(dataset_id=v.dataset_id)
        ]

        return {'locs': locs}


class ReferencesListView(ListView):
    """Show a list of all added research paper."""

    model = LocationInformation

    def get_template_names(self):
        return 'isodatabank/references.html'

    def get_context_data(self, **kwargs):
        validated = ValidationInformation.objects.filter(validated=True)
        locs = [
            l for v in validated \
                for l in LocationInformation.objects.filter(dataset_id=v.dataset_id)
        ]

        return {'locs': locs}


def download_view(request, dataset_id):
    """Generate an Excel file for download using the given dataset_id."""

    dataset = get_object_or_404(TiliaExcelFile, dataset_id=dataset_id)
    data = loads(dataset.data)

    with NamedTemporaryFile() as download_file:
        data.save(download_file.name)
        
        resp = HttpResponse(
            download_file,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(dataset_id)

        return resp
